from openpyxl import load_workbook

wb = load_workbook("student.xlsx")
sheet = wb.active

for row in sheet.iter_rows(values_only=True):
    print(row)